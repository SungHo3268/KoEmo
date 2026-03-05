"""
KoEmo 크라우드소싱 템플릿 XLSX 생성
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
EXAMPLE_FONT = Font(size=11, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_example(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER


def create_template(output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "KoEmo"

    headers = ["번호", "도메인", "카테고리", "단어군", "정답", "상황"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    examples = [
        ["(예시)", "감각 표현", "미각", "맵다, 얼큰하다, 알싸하다, 칼칼하다", "얼큰하다",
         "추운 겨울날 순대국밥 국물을 한 숟갈 떠먹으니 속이 확 풀리는 것이 국물이 참 ___."],
        ["(예시)", "감각 표현", "미각", "맵다, 얼큰하다, 알싸하다, 칼칼하다", "알싸하다",
         "초밥 위 와사비를 너무 많이 찍어 먹었더니 코끝이 찡하고 눈물이 핑 돌았다. 와사비가 ___."],
        ["(예시)", "감각 표현", "미각", "맵다, 얼큰하다, 알싸하다, 칼칼하다", "맵다",
         "라면에 고춧가루를 잔뜩 넣었더니 그냥 ___ 먹기 힘들었다."],
        ["(예시)", "감각 표현", "미각", "맵다, 얼큰하다, 알싸하다, 칼칼하다", "칼칼하다",
         "김치찌개에 청양고추를 넣었더니 국물 넘길 때 목구멍이 얼얼하게 ___."],
    ]
    for ex in examples:
        ws.append(ex)
        style_example(ws, ws.max_row, len(headers))

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 80

    wb.save(output_path)
    print(f"템플릿 생성: {output_path}")


if __name__ == "__main__":
    create_template("docs/KoEmo_template.xlsx")
