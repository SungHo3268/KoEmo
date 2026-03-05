"""
크라우드소싱 제출물(XLSX)을 JSONL로 변환하는 도구
단일 시트(도메인, 카테고리, 단어군, 정답, 상황)를 읽어 변환

사용법:
  python convert_xlsx_to_jsonl.py input.xlsx -o output.jsonl
"""

import json
import argparse
import sys
from openpyxl import load_workbook


DOMAIN_CODE_MAP = {
    "감각 표현": "sensory",
    "감정 표현": "emotional",
    "판단 표현": "judgement",
    "감상 표현": "appreciation",
    "상징 표현": "symbolic",
}

CATEGORY_CODE_MAP = {
    "시각": "visual", "미각": "gustatory", "촉각": "tactile",
    "청각": "auditory", "후각": "olfactory",
    "긍정 감정": "positive", "부정 감정": "negative", "복합 감정": "complex",
    "사회적 관계": "social", "능력/성품 판단": "ability",
    "능력/성품": "ability", "상황 판단": "situation",
    "심미적 평가": "aesthetic", "가치 평가": "value",
    "의태어": "mimetic", "의성어": "onomatopoeia",
}


def convert(input_path: str, output_path: str):
    wb = load_workbook(input_path, read_only=True)
    ws = wb.worksheets[0]

    items = []
    errors = []
    id_counters = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        domain = str(row[0]).strip() if row[0] else None
        category = str(row[1]).strip() if row[1] else None
        word_group_raw = str(row[2]).strip() if row[2] else None
        answer = str(row[3]).strip() if row[3] else None
        scenario = str(row[4]).strip() if row[4] else None

        if not all([domain, category, word_group_raw, answer, scenario]):
            continue

        words = [w.strip() for w in word_group_raw.split(",") if w.strip()]
        if len(words) < 2:
            errors.append(f"행 {row_num}: 단어군에 단어가 2개 미만")
            continue

        if "___" not in scenario:
            errors.append(f"행 {row_num}: 빈칸(___)이 없음")
            continue

        # ID 생성
        domain_code = DOMAIN_CODE_MAP.get(domain, domain)
        cat_code = CATEGORY_CODE_MAP.get(category, category)
        counter_key = f"{domain_code}_{cat_code}"
        id_counters[counter_key] = id_counters.get(counter_key, 0) + 1
        item_id = f"{domain_code}_{cat_code}_{id_counters[counter_key]:03d}"

        items.append({
            "id": item_id,
            "domain": domain,
            "category": category,
            "word_group": words,
            "scenario": scenario,
            "choices": words,
            "answer": answer,
        })

    wb.close()

    if errors:
        print(f"경고: {len(errors)}건의 오류", file=sys.stderr)
        for err in errors:
            print(f"  - {err}", file=sys.stderr)

    with open(output_path, "w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    print(f"변환 완료: {len(items)}문항 -> {output_path}")


def main():
    parser = argparse.ArgumentParser(description="XLSX -> JSONL 변환")
    parser.add_argument("input", help="입력 XLSX 파일")
    parser.add_argument("-o", "--output", default="output.jsonl", help="출력 JSONL 파일")
    args = parser.parse_args()
    convert(args.input, args.output)


if __name__ == "__main__":
    main()
