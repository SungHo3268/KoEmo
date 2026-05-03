"""
KoNuance Benchmark Evaluation Script
한국어 맥락 의존적 어휘 선택 능력 평가 도구

지원 입력: .xlsx (samples.xlsx) 또는 .jsonl (samples.jsonl)
지원 API:  openai, anthropic, vllm (OpenAI-compatible)

평가 모드:
  - individual: 문항별 독립 평가
  - set: 같은 유의어군 문항을 세트로 묶어 평가
  - sequential: 순차 제거 방식 (소거법 효과 측정)
  - both: individual + set (기본값)
  - all: individual + set + sequential

메트릭 (4가지 + sequential):
  1) Individual Item Acc: 개별 모드에서 문항별 정답률
  2) Individual Group Acc: 개별 모드에서 그룹 내 전체 정답 시 정답 처리
  3) Set Item Acc: 세트 모드에서 문항별 정답률
  4) Set Group Acc: 세트 모드에서 그룹 전체 정답률
  5) Sequential Round 1/2/3 Acc: 순차 제거 라운드별 정답률 (다중 시드 평균)
  6) Sequential Item Acc: 순차 제거 전체 정답률 (Round 4 자동 제외)

사용법:
  python evaluate.py --data data/samples.xlsx --provider openai --model gpt-4o
  python evaluate.py --data data/samples.xlsx --provider anthropic --model claude-sonnet-4-20250514
  python evaluate.py --data data/samples.xlsx --provider vllm --model meta-llama/Llama-3-8B --base-url http://localhost:8000/v1
  python evaluate.py --data data/samples.xlsx --provider openai --model gpt-4o --mode individual
  python evaluate.py --data data/samples.xlsx --provider openai --model gpt-4o --limit 10
"""

import json
import argparse
import os
import random
import re
import time
from pathlib import Path
from collections import defaultdict


# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------

def load_from_jsonl(path: str) -> list[dict]:
    items = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            item = json.loads(line)
            # Normalize: ensure choices field exists
            if "choices" not in item:
                item["choices"] = list(item["word_group"])
            # Normalize scenario marker
            item["scenario"] = item["scenario"].replace("___", "[정답]")
            items.append(item)
    return items


def load_from_xlsx(path: str) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    ws = wb.worksheets[0]

    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Detect column layout
    # Expected: 번호, 도메인, 카테고리, 유의어군, 정답, 상황
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.replace(" ", "")
        if h_lower in ("번호", "no", "번호(no)"):
            col_map["no"] = i
        elif "도메인" in h_lower:
            col_map["domain"] = i
        elif "카테고리" in h_lower:
            col_map["category"] = i
        elif "유의어군" in h_lower or "단어군" in h_lower:
            col_map["word_group"] = i
        elif "정답" in h_lower:
            col_map["answer"] = i
        elif "상황" in h_lower:
            col_map["scenario"] = i

    required = {"domain", "category", "word_group", "answer", "scenario"}
    missing = required - set(col_map.keys())
    if missing:
        raise ValueError(f"필수 열이 없습니다: {missing}. 헤더: {headers}")

    # Domain/category code maps for ID generation
    domain_code = {
        "감각 표현": "sensory", "감정 표현": "emotional", "판단 표현": "judgement",
        "감상 표현": "appreciation", "상징 표현": "symbolic",
    }
    category_code = {
        "시각": "visual", "미각": "gustatory", "촉각": "tactile",
        "청각": "auditory", "후각": "olfactory",
        "긍정 감정": "positive", "부정 감정": "negative", "복합 감정": "complex",
        "사회적 관계": "social", "능력/성품 판단": "ability", "능력/성품": "ability",
        "상황 판단": "situation", "상황/정도": "situation",
        "심미적 평가": "aesthetic", "가치 평가": "value",
        "의태어": "mimetic", "의성어": "onomatopoeia",
    }

    items = []
    id_counters: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)

        domain = str(row[col_map["domain"]]).strip() if row[col_map["domain"]] else ""
        category = str(row[col_map["category"]]).strip() if row[col_map["category"]] else ""
        wg_raw = str(row[col_map["word_group"]]).strip() if row[col_map["word_group"]] else ""
        answer = str(row[col_map["answer"]]).strip() if row[col_map["answer"]] else ""
        scenario = str(row[col_map["scenario"]]).strip() if row[col_map["scenario"]] else ""

        # Skip empty / example rows
        if not all([domain, category, wg_raw, answer, scenario]):
            continue
        if domain in ("None", "(예시)") or "(예시)" in str(row[col_map.get("no", 0)] or ""):
            continue

        words = [w.strip() for w in wg_raw.split(",") if w.strip()]
        if len(words) < 2:
            continue

        # Normalize scenario marker
        scenario = scenario.replace("___", "[정답]")

        # Generate ID
        dc = domain_code.get(domain, domain)
        cc = category_code.get(category, category)
        counter_key = f"{dc}_{cc}"
        id_counters[counter_key] = id_counters.get(counter_key, 0) + 1
        item_id = f"{counter_key}_{id_counters[counter_key]:03d}"

        items.append({
            "id": item_id,
            "domain": domain,
            "category": category,
            "word_group": words,
            "choices": list(words),
            "answer": answer,
            "scenario": scenario,
        })

    wb.close()
    return items


def load_benchmark(path: str) -> list[dict]:
    if path.endswith(".xlsx"):
        return load_from_xlsx(path)
    return load_from_jsonl(path)


# ---------------------------------------------------------------------------
# Grouping
# ---------------------------------------------------------------------------

def group_items(items: list[dict]) -> list[list[dict]]:
    """Group items by word_group (same synonym set)."""
    groups: dict[tuple, list[dict]] = {}
    for item in items:
        key = tuple(sorted(item["word_group"]))
        if key not in groups:
            groups[key] = []
        groups[key].append(item)
    return list(groups.values())


# ---------------------------------------------------------------------------
# Prompt Construction
# ---------------------------------------------------------------------------

def format_individual_prompt(item: dict, shuffle_seed: int | None = None) -> tuple[str, list[str]]:
    """Build individual-mode prompt. Returns (prompt_text, ordered_choices)."""
    choices = list(item["choices"])
    if shuffle_seed is not None:
        rng = random.Random(shuffle_seed)
        rng.shuffle(choices)

    display_scenario = item["scenario"].replace("[정답]", "___")

    n = len(choices)
    labels = [chr(65 + i) for i in range(n)]
    choices_text = "\n".join(f"  {labels[i]}. {choices[i]}" for i in range(n))
    labels_str = ", ".join(labels)

    prompt = (
        f"다음 상황에서 빈칸(___) 에 가장 자연스러운 표현을 하나만 골라 "
        f"{labels_str} 중 하나로 답하세요.\n"
        f"반드시 알파벳 한 글자로만 답하세요.\n\n"
        f"상황: {display_scenario}\n\n"
        f"선택지:\n{choices_text}\n\n"
        f"답:"
    )
    return prompt, choices


def format_set_prompt(group: list[dict], shuffle_seed: int | None = None) -> tuple[str, list[str], list[dict]]:
    """Build set-mode prompt. Returns (prompt_text, ordered_choices, ordered_items).

    All items in the group share the same word_group.
    Choices and item order are shuffled to prevent position bias.
    """
    items = list(group)
    choices = list(group[0]["choices"])

    if shuffle_seed is not None:
        rng = random.Random(shuffle_seed)
        rng.shuffle(choices)
        rng.shuffle(items)

    n = len(choices)
    labels = [chr(65 + i) for i in range(n)]
    choices_text = "\n".join(f"  {labels[i]}. {choices[i]}" for i in range(n))
    labels_str = ", ".join(labels)

    scenarios_lines = []
    for j, item in enumerate(items):
        display = item["scenario"].replace("[정답]", "___")
        scenarios_lines.append(f"상황 {j + 1}: {display}")
    scenarios_text = "\n".join(scenarios_lines)

    example = ", ".join(f"{j + 1}:{labels[j]}" for j in range(n))

    prompt = (
        f"다음은 같은 유의어군에 속하는 단어들을 사용하는 여러 상황입니다.\n"
        f"각 상황의 빈칸(___) 에 가장 자연스러운 표현을 하나씩 골라주세요.\n"
        f"단, 각 단어는 한 번만 사용할 수 있습니다.\n\n"
        f"{scenarios_text}\n\n"
        f"선택지:\n{choices_text}\n\n"
        f"각 상황에 대해 {labels_str} 중 하나씩 답하세요. 같은 알파벳을 중복 사용할 수 없습니다.\n"
        f"반드시 \"{example}\" 형식으로만 답하세요.\n\n"
        f"답:"
    )
    return prompt, choices, items


# ---------------------------------------------------------------------------
# Answer Parsing
# ---------------------------------------------------------------------------

def parse_individual_answer(response: str, choices: list[str]) -> str | None:
    response = response.strip()

    # Match A/B/C/D (or more) label
    match = re.search(r'\b([A-Za-z])\b', response)
    if match:
        idx = ord(match.group(1).upper()) - 65
        if 0 <= idx < len(choices):
            return choices[idx]

    # Direct text match
    for choice in choices:
        if choice in response:
            return choice

    return None


def parse_set_answer(response: str, choices: list[str], num_items: int) -> dict[int, str]:
    """Parse set-mode response like '1:A, 2:B, 3:C, 4:D'.

    Returns dict mapping item index (0-based) -> chosen word.
    """
    response = response.strip()
    results = {}

    matches = re.findall(r'(\d+)\s*:\s*([A-Za-z])', response)
    for num_str, label in matches:
        idx = int(num_str) - 1
        choice_idx = ord(label.upper()) - 65
        if 0 <= idx < num_items and 0 <= choice_idx < len(choices):
            results[idx] = choices[choice_idx]

    return results


# ---------------------------------------------------------------------------
# API Backends
# ---------------------------------------------------------------------------

SYSTEM_MSG_INDIVIDUAL = "당신은 한국어 전문가입니다. 질문에 대해 알파벳 한 글자로만 답하세요."
SYSTEM_MSG_SET = "당신은 한국어 전문가입니다. 질문에 대해 지정된 형식으로만 답하세요."


def _is_reasoning_model(model: str) -> bool:
    return model.startswith(("o1", "o3", "o4"))


def _needs_max_completion_tokens(model: str) -> bool:
    return model.startswith(("o1", "o3", "o4", "gpt-5"))


def call_openai(client, model: str, prompt: str, system_msg: str,
                is_reasoning: bool, max_tokens: int = 16) -> str:
    if is_reasoning:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": system_msg + "\n\n" + prompt}],
            max_completion_tokens=4096,
        )
    elif _needs_max_completion_tokens(model):
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": prompt},
            ],
            temperature=1,
            max_completion_tokens=4096,
        )
    else:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": prompt},
            ],
            temperature=0,
            max_tokens=max_tokens,
        )
    content = resp.choices[0].message.content
    return content.strip() if content else ""


def call_anthropic(client, model: str, prompt: str, system_msg: str,
                   max_tokens: int = 16) -> str:
    resp = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        temperature=0,
        system=system_msg,
        messages=[{"role": "user", "content": prompt}],
    )
    return resp.content[0].text.strip()


# ---------------------------------------------------------------------------
# Evaluation: Individual Mode
# ---------------------------------------------------------------------------

def evaluate_individual(
    items: list[dict],
    model: str,
    provider: str,
    api_key: str | None,
    base_url: str | None = None,
    shuffle_seed: int = 42,
) -> list[dict]:
    if provider in ("openai", "vllm"):
        from openai import OpenAI
        kwargs = {"api_key": api_key}
        if base_url:
            kwargs["base_url"] = base_url
        client = OpenAI(**kwargs)
        is_reasoning = _is_reasoning_model(model)
    elif provider == "anthropic":
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
    else:
        raise ValueError(f"지원하지 않는 provider: {provider}")

    results = []
    for i, item in enumerate(items):
        prompt, ordered_choices = format_individual_prompt(item, shuffle_seed=shuffle_seed)
        correct_answer = item["answer"]

        try:
            if provider in ("openai", "vllm"):
                raw = call_openai(client, model, prompt, SYSTEM_MSG_INDIVIDUAL,
                                  is_reasoning=(provider == "openai" and is_reasoning))
            else:
                raw = call_anthropic(client, model, prompt, SYSTEM_MSG_INDIVIDUAL)

            parsed = parse_individual_answer(raw, ordered_choices)
            is_correct = parsed == correct_answer

            if parsed is None:
                status = "PARSE_FAIL"
            elif is_correct:
                status = "O"
            else:
                status = "X"

            results.append({
                "id": item["id"],
                "domain": item["domain"],
                "category": item["category"],
                "group_key": ",".join(sorted(item["word_group"])),
                "model_response": raw,
                "parsed_answer": parsed,
                "correct_answer": correct_answer,
                "is_correct": is_correct,
            })
            print(f"  [{i+1:3d}/{len(items)}] {item['id']:<30s} {status:>10s}  (model={parsed}, gold={correct_answer})")

        except Exception as e:
            print(f"  [{i+1:3d}/{len(items)}] {item['id']:<30s}      ERROR  {e}")
            results.append({
                "id": item["id"],
                "domain": item["domain"],
                "category": item["category"],
                "group_key": ",".join(sorted(item["word_group"])),
                "model_response": str(e),
                "parsed_answer": None,
                "correct_answer": correct_answer,
                "is_correct": False,
            })

        time.sleep(0.3)

    return results


# ---------------------------------------------------------------------------
# Evaluation: Set Mode
# ---------------------------------------------------------------------------

def evaluate_set(
    items: list[dict],
    model: str,
    provider: str,
    api_key: str | None,
    base_url: str | None = None,
    shuffle_seed: int = 42,
) -> list[dict]:
    if provider in ("openai", "vllm"):
        from openai import OpenAI
        kwargs = {"api_key": api_key}
        if base_url:
            kwargs["base_url"] = base_url
        client = OpenAI(**kwargs)
        is_reasoning = _is_reasoning_model(model)
    elif provider == "anthropic":
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
    else:
        raise ValueError(f"지원하지 않는 provider: {provider}")

    groups = group_items(items)
    results = []

    for gi, group in enumerate(groups):
        prompt, ordered_choices, ordered_items = format_set_prompt(group, shuffle_seed=shuffle_seed)
        group_key = ",".join(sorted(group[0]["word_group"]))

        try:
            if provider in ("openai", "vllm"):
                raw = call_openai(client, model, prompt, SYSTEM_MSG_SET,
                                  is_reasoning=(provider == "openai" and is_reasoning),
                                  max_tokens=128)
            else:
                raw = call_anthropic(client, model, prompt, SYSTEM_MSG_SET,
                                     max_tokens=128)

            parsed_map = parse_set_answer(raw, ordered_choices, len(ordered_items))

            for j, item in enumerate(ordered_items):
                model_answer = parsed_map.get(j)
                correct_answer = item["answer"]
                is_correct = model_answer == correct_answer

                if model_answer is None:
                    status = "PARSE_FAIL"
                elif is_correct:
                    status = "O"
                else:
                    status = "X"

                results.append({
                    "id": item["id"],
                    "domain": item["domain"],
                    "category": item["category"],
                    "group_key": group_key,
                    "model_response": raw,
                    "parsed_answer": model_answer,
                    "correct_answer": correct_answer,
                    "is_correct": is_correct,
                })

            # Print group summary
            group_correct = sum(1 for j in range(len(ordered_items)) if parsed_map.get(j) == ordered_items[j]["answer"])
            group_total = len(ordered_items)
            group_status = "O" if group_correct == group_total else f"{group_correct}/{group_total}"
            print(f"  [Group {gi+1:3d}/{len(groups)}] {group_key:<40s} {group_status:>10s}")

        except Exception as e:
            print(f"  [Group {gi+1:3d}/{len(groups)}] {group_key:<40s}      ERROR  {e}")
            for item in group:
                results.append({
                    "id": item["id"],
                    "domain": item["domain"],
                    "category": item["category"],
                    "group_key": group_key,
                    "model_response": str(e),
                    "parsed_answer": None,
                    "correct_answer": item["answer"],
                    "is_correct": False,
                })

        time.sleep(0.3)

    return results


# ---------------------------------------------------------------------------
# Evaluation: Sequential Elimination Mode
# ---------------------------------------------------------------------------

def _init_client(provider, api_key, base_url=None):
    """Initialize API client. Returns (client, is_reasoning_flag)."""
    if provider in ("openai", "vllm"):
        from openai import OpenAI
        kwargs = {"api_key": api_key}
        if base_url:
            kwargs["base_url"] = base_url
        return OpenAI(**kwargs), True
    elif provider == "anthropic":
        import anthropic
        return anthropic.Anthropic(api_key=api_key), False
    else:
        raise ValueError(f"지원하지 않는 provider: {provider}")


def evaluate_sequential(
    items: list[dict],
    model: str,
    provider: str,
    api_key: str | None,
    base_url: str | None = None,
    shuffle_seeds: list[int] | None = None,
) -> dict:
    """Sequential elimination evaluation.

    For each group, questions are presented one at a time.
    After each answer, the chosen word is removed from the pool.
    Round 4 (last question, 1 choice left) is auto-assigned and excluded from scoring.

    Multiple seeds are used to average out ordering effects.

    Returns dict with per-seed details and averaged metrics.
    """
    if shuffle_seeds is None:
        shuffle_seeds = [42, 43, 44]

    client, is_openai = _init_client(provider, api_key, base_url)
    is_reasoning = is_openai and _is_reasoning_model(model)

    groups = group_items(items)
    all_seed_results = []

    for seed_idx, seed in enumerate(shuffle_seeds):
        print(f"\n  --- Seed {seed} ({seed_idx + 1}/{len(shuffle_seeds)}) ---")
        seed_results = []

        for gi, group in enumerate(groups):
            rng = random.Random(seed)
            ordered_items = list(group)
            rng.shuffle(ordered_items)

            remaining_choices = list(group[0]["choices"])
            rng.shuffle(remaining_choices)

            group_key = ",".join(sorted(group[0]["word_group"]))

            for round_num in range(len(ordered_items)):
                item = ordered_items[round_num]
                correct_answer = item["answer"]
                num_choices = len(remaining_choices)

                if round_num == len(ordered_items) - 1:
                    # Last round: auto-assign
                    model_answer = remaining_choices[0] if remaining_choices else None
                    is_correct = model_answer == correct_answer
                    status = "AUTO_O" if is_correct else "AUTO_X"

                    seed_results.append({
                        "id": item["id"],
                        "domain": item["domain"],
                        "category": item["category"],
                        "group_key": group_key,
                        "round": round_num + 1,
                        "num_choices": num_choices,
                        "auto": True,
                        "model_response": "(auto)",
                        "parsed_answer": model_answer,
                        "correct_answer": correct_answer,
                        "is_correct": is_correct,
                    })
                    print(f"  [G{gi+1:>2d} R{round_num+1}] {item['id']:<30s} {status:>10s}  (auto={model_answer}, gold={correct_answer})")
                    continue

                # Build prompt with remaining choices
                display_scenario = item["scenario"].replace("[정답]", "___")
                n = len(remaining_choices)
                labels = [chr(65 + i) for i in range(n)]
                choices_text = "\n".join(f"  {labels[i]}. {remaining_choices[i]}" for i in range(n))
                labels_str = ", ".join(labels)

                prompt = (
                    f"다음 상황에서 빈칸(___) 에 가장 자연스러운 표현을 하나만 골라 "
                    f"{labels_str} 중 하나로 답하세요.\n"
                    f"반드시 알파벳 한 글자로만 답하세요.\n\n"
                    f"상황: {display_scenario}\n\n"
                    f"선택지:\n{choices_text}\n\n"
                    f"답:"
                )

                try:
                    if provider in ("openai", "vllm"):
                        raw = call_openai(client, model, prompt, SYSTEM_MSG_INDIVIDUAL,
                                          is_reasoning=(provider == "openai" and is_reasoning))
                    else:
                        raw = call_anthropic(client, model, prompt, SYSTEM_MSG_INDIVIDUAL)

                    parsed = parse_individual_answer(raw, remaining_choices)
                    is_correct = parsed == correct_answer

                    if parsed is None:
                        status = "PARSE_FAIL"
                    elif is_correct:
                        status = "O"
                    else:
                        status = "X"

                    seed_results.append({
                        "id": item["id"],
                        "domain": item["domain"],
                        "category": item["category"],
                        "group_key": group_key,
                        "round": round_num + 1,
                        "num_choices": num_choices,
                        "auto": False,
                        "model_response": raw,
                        "parsed_answer": parsed,
                        "correct_answer": correct_answer,
                        "is_correct": is_correct,
                    })
                    print(f"  [G{gi+1:>2d} R{round_num+1}] {item['id']:<30s} {status:>10s}  (model={parsed}, gold={correct_answer}, choices={n})")

                    # Remove chosen word from pool
                    if parsed is not None and parsed in remaining_choices:
                        remaining_choices.remove(parsed)

                except Exception as e:
                    print(f"  [G{gi+1:>2d} R{round_num+1}] {item['id']:<30s}      ERROR  {e}")
                    seed_results.append({
                        "id": item["id"],
                        "domain": item["domain"],
                        "category": item["category"],
                        "group_key": group_key,
                        "round": round_num + 1,
                        "num_choices": num_choices,
                        "auto": False,
                        "model_response": str(e),
                        "parsed_answer": None,
                        "correct_answer": correct_answer,
                        "is_correct": False,
                    })

                time.sleep(0.3)

        all_seed_results.append(seed_results)

    return {"seeds": shuffle_seeds, "per_seed": all_seed_results}


def compute_sequential_metrics(seq_data: dict) -> dict:
    """Compute averaged metrics across seeds for sequential mode."""
    seeds = seq_data["seeds"]
    per_seed = seq_data["per_seed"]

    seed_metrics = []
    for seed_results in per_seed:
        # Item accuracy (excluding auto/Round 4)
        non_auto = [r for r in seed_results if not r["auto"]]
        item_correct = sum(1 for r in non_auto if r["is_correct"])
        item_total = len(non_auto)

        # Per-round accuracy
        round_stats = defaultdict(lambda: {"total": 0, "correct": 0})
        for r in seed_results:
            rd = r["round"]
            round_stats[rd]["total"] += 1
            if r["is_correct"]:
                round_stats[rd]["correct"] += 1

        # Group accuracy (all 4 rounds including auto must be correct)
        group_data = defaultdict(lambda: {"all_correct": True})
        for r in seed_results:
            if not r["is_correct"]:
                group_data[r["group_key"]]["all_correct"] = False
        num_groups = len(group_data)
        groups_correct = sum(1 for g in group_data.values() if g["all_correct"])

        seed_metrics.append({
            "item_accuracy": item_correct / item_total if item_total > 0 else 0,
            "item_correct": item_correct,
            "item_total": item_total,
            "group_accuracy": groups_correct / num_groups if num_groups > 0 else 0,
            "group_correct": groups_correct,
            "group_total": num_groups,
            "by_round": {
                rd: {
                    "total": s["total"],
                    "correct": s["correct"],
                    "accuracy": s["correct"] / s["total"] if s["total"] > 0 else 0,
                }
                for rd, s in sorted(round_stats.items())
            },
        })

    # Average across seeds
    avg_item_acc = sum(m["item_accuracy"] for m in seed_metrics) / len(seed_metrics)
    avg_group_acc = sum(m["group_accuracy"] for m in seed_metrics) / len(seed_metrics)

    all_rounds = set()
    for m in seed_metrics:
        all_rounds.update(m["by_round"].keys())

    avg_by_round = {}
    for rd in sorted(all_rounds):
        accs = [m["by_round"][rd]["accuracy"] for m in seed_metrics if rd in m["by_round"]]
        avg_by_round[rd] = sum(accs) / len(accs) if accs else 0

    return {
        "seeds": seeds,
        "averaged": {
            "item_accuracy": avg_item_acc,
            "group_accuracy": avg_group_acc,
            "by_round": avg_by_round,
        },
        "per_seed": seed_metrics,
    }


# ---------------------------------------------------------------------------
# Metrics & Reporting
# ---------------------------------------------------------------------------

def compute_metrics(results: list[dict]) -> dict:
    """Compute item-level and group-level accuracy from a list of results."""
    total = len(results)
    correct = sum(1 for r in results if r["is_correct"])
    no_answer = sum(1 for r in results if r["parsed_answer"] is None)

    # Domain / category breakdown (item-level)
    domain_stats = defaultdict(lambda: {"total": 0, "correct": 0})
    category_stats = defaultdict(lambda: {"total": 0, "correct": 0})

    for r in results:
        d = r["domain"]
        c = f"{d} > {r['category']}"
        domain_stats[d]["total"] += 1
        category_stats[c]["total"] += 1
        if r["is_correct"]:
            domain_stats[d]["correct"] += 1
            category_stats[c]["correct"] += 1

    # Group-level accuracy: all items in group must be correct
    group_results = defaultdict(lambda: {"total": 0, "correct": 0, "all_correct": True})
    for r in results:
        gk = r["group_key"]
        group_results[gk]["total"] += 1
        if r["is_correct"]:
            group_results[gk]["correct"] += 1
        else:
            group_results[gk]["all_correct"] = False

    num_groups = len(group_results)
    groups_all_correct = sum(1 for g in group_results.values() if g["all_correct"])

    def acc(s):
        return s["correct"] / s["total"] if s["total"] > 0 else 0

    return {
        "item": {
            "total": total,
            "correct": correct,
            "accuracy": correct / total if total > 0 else 0,
            "no_answer": no_answer,
        },
        "group": {
            "total": num_groups,
            "correct": groups_all_correct,
            "accuracy": groups_all_correct / num_groups if num_groups > 0 else 0,
        },
        "by_domain": {
            d: {"total": s["total"], "correct": s["correct"], "accuracy": acc(s)}
            for d, s in sorted(domain_stats.items())
        },
        "by_category": {
            c: {"total": s["total"], "correct": s["correct"], "accuracy": acc(s)}
            for c, s in sorted(category_stats.items())
        },
    }


def print_report(individual_metrics: dict | None, set_metrics: dict | None,
                 sequential_metrics: dict | None, model_name: str):
    print()
    print("=" * 68)
    print(f"  KoNuance Benchmark Results - {model_name}")
    print("=" * 68)

    sections = []
    if individual_metrics:
        sections.append(("Individual Mode", individual_metrics))
    if set_metrics:
        sections.append(("Set Mode", set_metrics))

    for title, metrics in sections:
        item = metrics["item"]
        group = metrics["group"]

        print(f"\n  [{title}]")
        print(f"  {'─' * 60}")
        print(f"  Item Accuracy:   {item['correct']:>4d}/{item['total']:<4d} ({item['accuracy']:.1%})")
        print(f"  Group Accuracy:  {group['correct']:>4d}/{group['total']:<4d} ({group['accuracy']:.1%})")
        if item["no_answer"] > 0:
            print(f"  Parse failures:  {item['no_answer']}")

        print(f"\n  {'Domain':<16} {'Correct':>7} {'Total':>6} {'Acc':>8}")
        print("  " + "-" * 40)
        for domain, s in metrics["by_domain"].items():
            print(f"  {domain:<14} {s['correct']:>7} {s['total']:>6} {s['accuracy']:>8.1%}")

        print(f"\n  {'Category':<32} {'Correct':>7} {'Total':>6} {'Acc':>8}")
        print("  " + "-" * 56)
        for cat, s in metrics["by_category"].items():
            print(f"  {cat:<30} {s['correct']:>7} {s['total']:>6} {s['accuracy']:>8.1%}")

    if sequential_metrics:
        avg = sequential_metrics["averaged"]
        seeds = sequential_metrics["seeds"]

        print(f"\n  [Sequential Elimination Mode]  (seeds: {seeds})")
        print(f"  {'─' * 60}")
        for rd, acc in avg["by_round"].items():
            n_choices = 4 - rd + 1
            label = f"auto" if n_choices == 1 else f"{n_choices} choices"
            print(f"  Round {rd} ({label:>10s}):  {acc:.1%}")
        print(f"  {'─' * 60}")
        print(f"  Item Accuracy (R1-R3):  {avg['item_accuracy']:.1%}")
        print(f"  Group Accuracy:         {avg['group_accuracy']:.1%}")

        # Per-seed breakdown
        print(f"\n  Per-seed breakdown:")
        for i, sm in enumerate(sequential_metrics["per_seed"]):
            seed = seeds[i]
            rounds_str = "  ".join(f"R{rd}:{s['accuracy']:.0%}" for rd, s in sm["by_round"].items())
            print(f"    seed={seed}:  Item={sm['item_accuracy']:.1%}  Group={sm['group_accuracy']:.1%}  {rounds_str}")

    print()
    print("=" * 68)


def save_results(
    individual_results: list[dict] | None,
    individual_metrics: dict | None,
    set_results: list[dict] | None,
    set_metrics: dict | None,
    sequential_data: dict | None,
    sequential_metrics: dict | None,
    model_name: str,
    output_dir: str,
):
    os.makedirs(output_dir, exist_ok=True)
    safe_name = re.sub(r'[/:\\]', '_', model_name)
    timestamp = time.strftime("%Y%m%d_%H%M%S")

    payload = {
        "model": model_name,
        "timestamp": timestamp,
    }

    if individual_metrics:
        payload["individual"] = {
            "metrics": individual_metrics,
            "details": individual_results,
        }
    if set_metrics:
        payload["set"] = {
            "metrics": set_metrics,
            "details": set_results,
        }
    if sequential_metrics:
        payload["sequential"] = {
            "metrics": sequential_metrics,
            "details": sequential_data["per_seed"] if sequential_data else [],
        }

    result_path = os.path.join(output_dir, f"{safe_name}_{timestamp}.json")
    with open(result_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"\n  Results saved: {result_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="KoNuance Benchmark 평가 도구",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python evaluate.py --data data/samples.xlsx --provider openai --model gpt-4o
  python evaluate.py --data data/samples.xlsx --provider anthropic --model claude-sonnet-4-20250514
  python evaluate.py --data data/samples.xlsx --provider vllm --model llama3 --base-url http://localhost:8000/v1
  python evaluate.py --data data/samples.xlsx --provider openai --model gpt-4o --mode set
  python evaluate.py --data data/samples.xlsx --provider openai --model gpt-4o --mode sequential
  python evaluate.py --data data/samples.xlsx --provider openai --model gpt-4o --mode all
        """,
    )
    parser.add_argument("--data", default="data/samples.xlsx", help="벤치마크 데이터 경로 (.xlsx 또는 .jsonl)")
    parser.add_argument("--provider", choices=["openai", "anthropic", "vllm"], required=True, help="API 제공자")
    parser.add_argument("--model", required=True, help="모델명")
    parser.add_argument("--mode", choices=["individual", "set", "sequential", "both", "all"], default="both",
                        help="평가 모드 (기본: both = individual+set)")
    parser.add_argument("--api-key", default=None, help="API 키 (미지정 시 환경변수)")
    parser.add_argument("--base-url", default=None, help="vLLM 등 커스텀 API base URL")
    parser.add_argument("--output-dir", default="results", help="결과 저장 디렉토리")
    parser.add_argument("--limit", type=int, default=None, help="평가 문항 수 제한")
    parser.add_argument("--seed", type=int, default=42, help="선택지 셔플 시드 (position bias 방지)")
    parser.add_argument("--seq-seeds", type=int, nargs="+", default=[42, 43, 44],
                        help="Sequential 모드 셔플 시드 목록 (기본: 42 43 44)")
    args = parser.parse_args()

    # Resolve API key: --api-key > env var > utils/*.txt
    api_key = args.api_key
    if not api_key:
        if args.provider in ("openai", "vllm"):
            api_key = os.environ.get("OPENAI_API_KEY")
        elif args.provider == "anthropic":
            api_key = os.environ.get("ANTHROPIC_API_KEY")

    if not api_key:
        key_files = {
            "openai": "utils/openai_api_key.txt",
            "vllm": "utils/openai_api_key.txt",
            "anthropic": "utils/anthropic_api_key.txt",
        }
        key_path = Path(__file__).parent / key_files.get(args.provider, "")
        if key_path.exists():
            api_key = key_path.read_text().strip()

    if not api_key and args.provider == "vllm":
        api_key = "EMPTY"

    if not api_key:
        env_var = "OPENAI_API_KEY" if args.provider == "openai" else "ANTHROPIC_API_KEY"
        print(f"ERROR: API 키가 필요합니다. --api-key, {env_var} 환경변수, 또는 utils/ 키 파일을 설정하세요.")
        return

    # Load data
    items = load_benchmark(args.data)
    if args.limit:
        items = items[: args.limit]

    groups = group_items(items)
    run_individual = args.mode in ("individual", "both", "all")
    run_set = args.mode in ("set", "both", "all")
    run_sequential = args.mode in ("sequential", "all")

    print(f"\n  KoNuance Benchmark Evaluation")
    print(f"  Model:    {args.model}")
    print(f"  Provider: {args.provider}")
    print(f"  Mode:     {args.mode}")
    print(f"  Items:    {len(items)}  ({len(groups)} groups)")
    print(f"  Data:     {args.data}")

    individual_results = None
    individual_metrics = None
    set_results = None
    set_metrics = None
    sequential_data = None
    sequential_metrics = None

    eval_kwargs = dict(
        items=items,
        model=args.model,
        provider=args.provider,
        api_key=api_key,
        base_url=args.base_url,
        shuffle_seed=args.seed,
    )

    # --- Individual Mode ---
    if run_individual:
        print(f"\n  {'─' * 60}")
        print(f"  [Individual Mode]")
        print(f"  {'─' * 60}")
        individual_results = evaluate_individual(**eval_kwargs)
        individual_metrics = compute_metrics(individual_results)

    # --- Set Mode ---
    if run_set:
        print(f"\n  {'─' * 60}")
        print(f"  [Set Mode]")
        print(f"  {'─' * 60}")
        set_results = evaluate_set(**eval_kwargs)
        set_metrics = compute_metrics(set_results)

    # --- Sequential Elimination Mode ---
    if run_sequential:
        print(f"\n  {'─' * 60}")
        print(f"  [Sequential Elimination Mode]  (seeds: {args.seq_seeds})")
        print(f"  {'─' * 60}")
        seq_kwargs = dict(eval_kwargs)
        del seq_kwargs["shuffle_seed"]
        sequential_data = evaluate_sequential(**seq_kwargs, shuffle_seeds=args.seq_seeds)
        sequential_metrics = compute_sequential_metrics(sequential_data)

    # Report
    print_report(individual_metrics, set_metrics, sequential_metrics, args.model)
    save_results(individual_results, individual_metrics,
                 set_results, set_metrics,
                 sequential_data, sequential_metrics,
                 args.model, args.output_dir)


if __name__ == "__main__":
    main()
